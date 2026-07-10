import base64
import json
import urllib.request
import urllib.error
import app as st
import os
import uuid
from typing import List

# Try to import openpyxl for xlsx handling; if unavailable, we'll disable XLSX features
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def rerun_app() -> None:
    """Safely trigger a Streamlit rerun across Streamlit versions.

    Tries `st.experimental_rerun()` first; if unavailable, attempts to raise
    the internal `RerunException` from known module paths. If neither is
    available, sets a session flag so the UI can indicate a refresh is needed.
    """
    try:
        if hasattr(st, "experimental_rerun"):
            st.experimental_rerun()
            return
    except Exception:
        pass

    # Try known internal locations for RerunException
    for modpath in (
        "streamlit.runtime.scriptrunner.script_runner",
        "streamlit.runtime.scriptrunner",
        "streamlit.scriptrunner.script_runner",
        "streamlit.ScriptRunner",
    ):
        try:
            mod = __import__(modpath, fromlist=["RerunException"])
            RerunException = getattr(mod, "RerunException")
            raise RerunException()
        except Exception:
            continue

    # Fallback: set a flag (UI can instruct user to refresh)
    st.session_state["_needs_rerun"] = True

OLLAMA_MODEL = "qwen2.5-coder:7b"
OLLAMA_BASE_URL = "http://localhost:11434"
AI_SYSTEM_PROMPT_DEFAULT = "You are a helpful and concise assistant. Answer clearly and thoughtfully. Use any of the provided sources to inform your answers, and cite them when relevant. If the user asks for information not covered by the sources, respond with 'I don't know.'"

st.set_page_config(page_title="Local Ollama Chat", page_icon="🤖")
st.write(f"Chat with a locally running Ollama model ({OLLAMA_MODEL}).")

# Sidebar: only show Sources management
with st.sidebar:
    st.header("Sources")
    if not OPENPYXL_AVAILABLE:
        st.error("openpyxl not available — install openpyxl to enable sources management.")

    SOURCES_FILE = "sources.xlsx"

    def ensure_sources_file(path: str) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sources"
            ws["A1"] = "Source"
            wb.save(path)

    def load_sources(path: str) -> List[str]:
        if not OPENPYXL_AVAILABLE:
            return []
        ensure_sources_file(path)
        wb = load_workbook(path)
        ws = wb.active
        sources = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val is None:
                continue
            sources.append(str(val))
        return sources

    def write_sources(path: str, sources: List[str]) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sources"
        ws["A1"] = "Source"
        for i, src in enumerate(sources, start=2):
            ws.cell(row=i, column=1, value=src)
        wb.save(path)

    # Initialize session state sources from file if needed
    def create_source_items(texts: List[str]) -> List[dict]:
        return [{"id": str(uuid.uuid4()), "text": text} for text in texts]

    if "sources" not in st.session_state:
        raw_sources = load_sources(SOURCES_FILE) if OPENPYXL_AVAILABLE else []
        st.session_state["sources"] = create_source_items(raw_sources)

    if "delete_source_id" not in st.session_state:
        st.session_state.delete_source_id = None

    new_src = st.text_area("Enter source text", key="new_source_text")
    if st.button("Add source"):
        if new_src and OPENPYXL_AVAILABLE:
            st.session_state["sources"].append({"id": str(uuid.uuid4()), "text": new_src.strip()})
            write_sources(SOURCES_FILE, [item["text"] for item in st.session_state["sources"]])

    st.markdown(
        "<style>"
        ".source-box { border: 1px solid #ddd; border-radius: 14px; padding: 12px; background: #fafafa; }"
        ".source-row { display: flex; align-items: center; justify-content: space-between; padding: 6px 0; }"
        ".source-row > div { width: auto; }"
        ".source-label { flex: 1; margin-right: 6px; }"
        ".trash-button button { background: transparent; border: none; color: #e03e2d; font-size: 18px; padding: 0 6px; margin-left: -6px; }"
        "</style>",
        unsafe_allow_html=True,
    )
    with st.expander("Source List", expanded=True):
        st.write("Select sources to include in the chat:")
        selected_sources = []
        for source in st.session_state["sources"]:
            key = f"src_chk_{source['id']}"
            cols = st.columns([0.85, 0.15])
            with cols[0]:
                checked = st.checkbox(source["text"], key=key)
            with cols[1]:
                if st.button("🗑️", key=f"del_{source['id']}", help="Delete this source"):
                    st.session_state.delete_source_id = source["id"]
            if checked:
                selected_sources.append(source["text"])
        st.markdown("</div>", unsafe_allow_html=True)

    if st.session_state.delete_source_id is not None:
        st.session_state["sources"] = [item for item in st.session_state["sources"] if item["id"] != st.session_state.delete_source_id]
        write_sources(SOURCES_FILE, [item["text"] for item in st.session_state["sources"]])
        checkbox_key = f"src_chk_{st.session_state.delete_source_id}"
        if checkbox_key in st.session_state:
            del st.session_state[checkbox_key]
        st.session_state.delete_source_id = None
        rerun_app()

    st.session_state.selected_sources = selected_sources

ollama_model = OLLAMA_MODEL
ollama_base_url = OLLAMA_BASE_URL

def get_ollama_candidates(base_url: str) -> list[str]:
    cleaned = base_url.strip().rstrip('/')
    candidates = [cleaned]

    if cleaned.startswith("http://localhost"):
        candidates.append(cleaned.replace("http://localhost", "http://127.0.0.1"))
    elif cleaned.startswith("http://127.0.0.1"):
        candidates.append(cleaned.replace("http://127.0.0.1", "http://localhost"))

    if not cleaned.endswith(":11434"):
        candidates.append(f"{cleaned}:11434")

    return list(dict.fromkeys(candidates))


def check_ollama_connection(base_url: str) -> str:
    for candidate in get_ollama_candidates(base_url):
        try:
            with urllib.request.urlopen(f"{candidate}/api/tags", timeout=10) as response:
                return "ok"
        except Exception as exc:
            last_error = exc
    return f"Could not reach the local Ollama server. Start Ollama with 'ollama serve' and confirm the base URL. Last error: {last_error}"


def query_ollama(prompt: str, base_url: str, model: str, system_prompt: str, image_bytes: bytes | None = None) -> str:
    image_b64 = base64.b64encode(image_bytes).decode("utf-8") if image_bytes else None
    system_message = {"role": "system", "content": system_prompt}

    if image_b64:
        payload = {
            "model": model,
            "messages": [
                system_message,
                {
                    "role": "user",
                    "content": prompt,
                    "images": [image_b64],
                }
            ],
            "stream": False,
        }
    else:
        payload = {
            "model": model,
            "messages": [system_message, {"role": "user", "content": prompt}],
            "stream": False,
        }

    last_error = None
    for candidate in get_ollama_candidates(base_url):
        request = urllib.request.Request(
            f"{candidate}/api/chat",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                data = json.loads(response.read().decode("utf-8"))
                return data.get("message", {}).get("content", "").strip()
        except Exception as exc:
            last_error = exc

    if isinstance(last_error, urllib.error.URLError):
        return f"Sorry, I could not reach your local Ollama server: {last_error}"
    return f"Sorry, something went wrong: {last_error}"



# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = [{"role": "assistant", "content": "What do you need help with?"}]

connection_status = check_ollama_connection(ollama_base_url)
if connection_status != "ok":
    st.warning(connection_status)
    st.caption("If Ollama is installed, open a terminal and run: ollama serve")

# Display chat messages from history on app rerun
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

# Accept user input (text-only)
prompt = st.chat_input("What is up?")
if prompt is not None:
    user_text = prompt.strip()
    if user_text:
        st.session_state.messages.append({"role": "user", "content": user_text})

        with st.chat_message("user"):
            st.markdown(user_text)

        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                selected_sources = st.session_state.get("selected_sources", [])
                if selected_sources:
                    combined_prompt = user_text + "\n\nIncluded sources:\n" + "\n\n".join(selected_sources)
                else:
                    combined_prompt = user_text

                assistant_response = query_ollama(
                    prompt=combined_prompt,
                    base_url=ollama_base_url,
                    model=ollama_model,
                    system_prompt=AI_SYSTEM_PROMPT_DEFAULT,
                )
            st.markdown(assistant_response)

        st.session_state.messages.append({"role": "assistant", "content": assistant_response})
